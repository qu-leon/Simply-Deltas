"""
Library functions to manipulate EMS flow compare excel files
"""

import sys
import os
import openpyxl
import win32com.client as win32
from pathlib import Path
import PySimpleGUI as sg

sg.theme("SystemDefault")


def resource_path(relative_path):
    """Get the absolute path to the resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


class ExcelCompare:
    """Class functions for handling EMS flow compare"""

    def __init__(self, save_file) -> None:
        self.save_file = save_file

    def compare_columns_and_generate_report(self, file_path) -> None:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        dlp1, dlp2 = self.get_sheet_title(ws)

        differences = []
        for row in range(25, ws.max_row + 1):
            d_val = ws[f"D{row}"].value
            t_val = ws[f"T{row}"].value
            e_val = ws[f"E{row}"].value
            u_val = ws[f"U{row}"].value
            f_val = ws[f"F{row}"].value
            g_val = ws[f"G{row}"].value
            h_val = ws[f"H{row}"].value
            i_val = ws[f"I{row}"].value
            k_val = ws[f"K{row}"].value
            v_val = ws[f"V{row}"].value
            w_val = ws[f"W{row}"].value
            x_val = ws[f"X{row}"].value
            y_val = ws[f"Y{row}"].value
            aa_val = ws[f"AA{row}"].value

            if d_val != t_val:
                differences.append(
                    (
                        d_val,
                        e_val,
                        f_val,
                        g_val,
                        h_val,
                        i_val,
                        k_val,
                        t_val,
                        u_val,
                        v_val,
                        w_val,
                        x_val,
                        y_val,
                        aa_val,
                    )
                )

        return differences, dlp1, dlp2

    def get_sheet_title(self, ws):
        sheet_title = ws.title.strip()
        parts = sheet_title.split()

        if len(parts) != 2:
            raise ValueError(
                f"Expected exactly two DLPs in sheet title, got: '{sheet_title}'"
            )

        return parts[0], parts[1]

    def create_outlook_draft(self, differences, file_name, dlp1, dlp2) -> None:
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)

        mail.Subject = f"Oper deltas found in: {file_name}"
        mail.To = ""

        if differences:
            html_body = f"""
            <html>
            <body>
            <p>The following deltas were found between <b>{dlp1}</b> and <b>{dlp2}</b>:</p>
            <table border="1" cellpadding="6" cellspacing="0" style="border-collapse: collapse; font-family: Arial; font-size: 12px;">
                <tr style="background-color: #f2f2f2;">
                    <th>OPER</th>
                    <th>INST</th>
                    <th>MER</th>
                    <th>SPL</th>
                    <th>SIF</th>
                    <th>SIF Value</th>
                    <th>OPER</th>
                    <th>INST</th>
                    <th>MER</th>
                    <th>SPL</th>
                    <th>SIF</th>
                    <th>SIF Value</th>
                </tr>
            """
            for (
                d_val,
                e_val,
                f_val,
                g_val,
                h_val,
                i_val,
                k_val,
                t_val,
                u_val,
                v_val,
                w_val,
                x_val,
                y_val,
                aa_val,
            ) in differences:
                html_body += f"""
                <tr>
                    <td>{d_val if d_val is not None else ''}{' '}{e_val if e_val is not None else ''}</td>
                    <td>{f_val if f_val is not None else ''}</td>
                    <td>{g_val if g_val is not None else ''}</td>
                    <td>{h_val if h_val is not None else ''}</td>
                    <td>{i_val if i_val is not None else ''}</td>
                    <td>{k_val if k_val is not None else ''}</td>
                    <td>{t_val if t_val is not None else ''}{' '}{u_val if u_val is not None else ''}</td>
                    <td>{v_val if v_val is not None else ''}</td>
                    <td>{w_val if w_val is not None else ''}</td>
                    <td>{x_val if x_val is not None else ''}</td>
                    <td>{y_val if y_val is not None else ''}</td>
                    <td>{aa_val if aa_val is not None else ''}</td>
                </tr>
                """
            html_body += """
            </table>
            </body>
            </html>
            """
        else:
            html_body = f"""<p>No differences were found between <b>{dlp1}</b> and <b>{dlp2}</b>.</p>"""

        mail.HTMLBody = html_body
        mail.Display()

    def validate_excel_format(self, ws) -> None:
        errors = []
        # Check if columns D and T are within range
        if ws.max_column < 20:
            errors.append("The sheet appears to be missing Lotplan opers.")
        # Check if rows from 25 exist
        if ws.max_row < 25:
            errors.append("The sheet does not have 25 or more rows.")

        return errors

    def run(self, file_path) -> None:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        format_errors = self.validate_excel_format(ws)
        if format_errors:
            sg.popup_error(
                "File validation failed:\n\n" + "\n".join(format_errors),
                title="Format Error",
                icon=resource_path("icons/compare.ico"),
            )
            return

        differences, dlp1, dlp2 = self.compare_columns_and_generate_report(file_path)
        self.create_outlook_draft(differences, Path(file_path).name, dlp1, dlp2)
